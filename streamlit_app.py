import io
import math
import os
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

INPUT_SHEET_MAIN = "Sheet1"
VENDOR_FALLBACK_COL = "Vendor"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


@dataclass
class PriorityRule:
    whse: str
    mode: str  # SI | SS
    value: float  # decimal


def normalize_whse(v) -> str:
    if pd.isna(v):
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
    except Exception:
        pass
    return str(v).strip()


def normalize_pct(value) -> float:
    if value is None:
        return 0.0
    value = float(value)
    if value > 1:
        value = value / 100.0
    return value


def round_to_int_units(value: float) -> int:
    return int(round(float(value)))


# SS% = Current SI / Average of SS Wk3

def safe_ss_ratio(current_si: float, ss_target: float) -> float:
    if ss_target <= 0:
        if current_si > 0:
            return math.inf
        if current_si < 0:
            return -math.inf
        return 0.0
    return current_si / ss_target


def ensure_vendor_col(df: pd.DataFrame) -> pd.DataFrame:
    vendor_cols = [c for c in df.columns if "vendor" in str(c).lower()]
    if not vendor_cols:
        df = df.copy()
        df[VENDOR_FALLBACK_COL] = ""
    return df


def load_input(file_obj) -> pd.DataFrame:
    df = pd.read_excel(file_obj, sheet_name=INPUT_SHEET_MAIN)
    df.columns = [str(c).strip() for c in df.columns]

    required = [
        "Item",
        "ProdResourceID",
        "Whse",
        "F Wk3",
        "Sum of SI Wk3",
        "Average of SS Wk3",
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Thiếu cột ở {INPUT_SHEET_MAIN}: {missing}")

    df = ensure_vendor_col(df)
    df = df.copy()
    df["Item"] = pd.to_numeric(df["Item"], errors="coerce")
    df["Whse"] = df["Whse"].map(normalize_whse)

    for c in ["F Wk3", "Sum of SI Wk3", "Average of SS Wk3"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    if "Sum of SI-SS Wk3" in df.columns:
        df["Sum of SI-SS Wk3"] = pd.to_numeric(df["Sum of SI-SS Wk3"], errors="coerce").fillna(0)
    else:
        df["Sum of SI-SS Wk3"] = pd.NA

    df["Current SI"] = df["Sum of SI Wk3"]
    df["Current SS%"] = df.apply(
        lambda r: safe_ss_ratio(float(r["Current SI"]), float(r["Average of SS Wk3"])), axis=1
    )

    totals = df.groupby("Item", as_index=False)["F Wk3"].sum().rename(columns={"F Wk3": "Firm PO Total"})
    df = df.merge(totals, on="Item", how="left")
    return df


def current_si_after(row: dict) -> int:
    return int(row["current_si"] + (row["final_f"] - row["orig_f"]))


def current_ss_after(row: dict) -> float:
    return safe_ss_ratio(current_si_after(row), row["ss_target"])


def compute_priority_target_final(row: dict, rule: PriorityRule) -> int:
    orig_f = row["orig_f"]
    current_si = row["current_si"]
    ss_target = row["ss_target"]

    if rule.mode == "SI":
        pct = max(0.0, min(1.0, float(rule.value)))
        target_si_after = current_si * (1.0 - pct)
        final_f = orig_f + (target_si_after - current_si)
        return max(0, round_to_int_units(final_f))

    if rule.mode == "SS":
        target_si_after = ss_target * float(rule.value)
        final_f = orig_f + (target_si_after - current_si)
        return max(0, round_to_int_units(final_f))

    return int(orig_f)


def build_rows(group: pd.DataFrame, item_rules: Dict[str, PriorityRule]) -> Tuple[List[dict], int]:
    rows: List[dict] = []
    for _, r in group.iterrows():
        row = {
            "item": r["Item"],
            "prod": r["ProdResourceID"],
            "whse": normalize_whse(r["Whse"]),
            "orig_f": round_to_int_units(r["F Wk3"]),
            "current_si": round_to_int_units(r["Current SI"]),
            "ss_target": float(r["Average of SS Wk3"]),
            "final_f": 0,
            "priority_rule_mode": "",
            "priority_rule_value": None,
            "priority_target_f_after": None,
        }
        rule = item_rules.get(row["whse"])
        if rule is not None:
            row["priority_rule_mode"] = rule.mode
            row["priority_rule_value"] = rule.value
            row["priority_target_f_after"] = compute_priority_target_final(row, rule)
        rows.append(row)
    total_f = round_to_int_units(group["Firm PO Total"].iloc[0])
    return rows, total_f


def choose_priority_recipient(rows: List[dict], priority_indices: List[int]) -> Optional[int]:
    candidates = []
    for idx in priority_indices:
        row = rows[idx]
        target = row.get("priority_target_f_after")
        if target is None:
            continue
        if row["final_f"] >= target:
            continue

        gap = target - row["final_f"]
        if row["priority_rule_mode"] == "SI":
            primary_metric = current_si_after(row)
            secondary_metric = current_ss_after(row)
        else:
            primary_metric = current_ss_after(row)
            secondary_metric = current_si_after(row)

        candidates.append((gap, primary_metric, secondary_metric, row["whse"], idx))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (-x[0], x[1], x[2], x[3]))
    return candidates[0][4]


def choose_phase1_recipient(rows: List[dict], candidate_indices: List[int]) -> Optional[int]:
    candidates = []
    for idx in candidate_indices:
        after_si = current_si_after(rows[idx])
        if after_si < 0:
            ratio = safe_ss_ratio(after_si, rows[idx]["ss_target"])
            candidates.append((after_si, ratio, rows[idx]["whse"], idx))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (x[0], x[1], x[2]))
    return candidates[0][3]


def choose_phase2_recipient(rows: List[dict], candidate_indices: List[int]) -> Optional[int]:
    candidates = []
    for idx in candidate_indices:
        after_si = current_si_after(rows[idx])
        ratio = safe_ss_ratio(after_si, rows[idx]["ss_target"])
        candidates.append((ratio, after_si, rows[idx]["whse"], idx))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (x[0], x[1], x[2]))
    return candidates[0][3]


def allocate_item(group: pd.DataFrame, item_rules: Dict[str, PriorityRule]) -> pd.DataFrame:
    rows, total_f = build_rows(group, item_rules)
    remaining = total_f

    priority_indices = [i for i, r in enumerate(rows) if r["priority_rule_mode"]]
    non_priority_indices = [i for i, r in enumerate(rows) if not r["priority_rule_mode"]]

    while remaining > 0:
        idx = choose_priority_recipient(rows, priority_indices)
        if idx is None:
            break
        rows[idx]["final_f"] += 1
        remaining -= 1

    allocation_pool = non_priority_indices[:] if non_priority_indices else priority_indices[:]

    while remaining > 0:
        idx = choose_phase1_recipient(rows, allocation_pool)
        if idx is None:
            break
        rows[idx]["final_f"] += 1
        remaining -= 1

    while remaining > 0:
        idx = choose_phase2_recipient(rows, allocation_pool)
        if idx is None:
            break
        rows[idx]["final_f"] += 1
        remaining -= 1

    out = group.copy().reset_index(drop=True)
    out["F Wk3 Original"] = out["F Wk3"].round().astype(int)
    out["F Wk3 After Destination Change"] = [r["final_f"] for r in rows]
    out["Net Destination Change"] = out["F Wk3 After Destination Change"] - out["F Wk3 Original"]
    out["Current SI After"] = out["Current SI"] + out["Net Destination Change"]
    out["SS % After"] = out.apply(
        lambda r: safe_ss_ratio(float(r["Current SI After"]), float(r["Average of SS Wk3"])), axis=1
    )
    out["Remaining Unallocated PO"] = total_f - int(out["F Wk3 After Destination Change"].sum())
    out["Priority Rule Mode"] = [r["priority_rule_mode"] for r in rows]
    out["Priority Rule Value"] = [r["priority_rule_value"] for r in rows]
    out["Priority Target F After"] = [r["priority_target_f_after"] for r in rows]

    before_total = int(out["F Wk3 Original"].sum())
    after_total = int(out["F Wk3 After Destination Change"].sum())
    if before_total != after_total:
        raise ValueError(
            f"Item {out['Item'].iloc[0]}: tổng firm không bảo toàn, before={before_total}, after={after_total}."
        )

    return out


def build_detail_output(detail: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "Item",
        "ProdResourceID",
        "Whse",
        "F Wk3 Original",
        "Current SI",
        "Average of SS Wk3",
        "Current SS%",
        "Firm PO Total",
        "F Wk3 After Destination Change",
        "Net Destination Change",
        "Current SI After",
        "SS % After",
        "Remaining Unallocated PO",
        "Priority Rule Mode",
        "Priority Rule Value",
        "Priority Target F After",
    ]
    vendor_cols = [c for c in detail.columns if "vendor" in str(c).lower() and c not in preferred]
    if not vendor_cols and VENDOR_FALLBACK_COL in detail.columns:
        vendor_cols = [VENDOR_FALLBACK_COL]
    final_cols = [c for c in preferred if c in detail.columns] + vendor_cols
    return detail[final_cols].copy()


def build_summary(detail: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, g in detail.groupby("Item", sort=True):
        rows.append(
            {
                "Item": g["Item"].iloc[0],
                "ProdResourceID": g["ProdResourceID"].iloc[0],
                "Firm PO Total": int(g["Firm PO Total"].iloc[0]),
                "Total F Before": int(g["F Wk3 Original"].sum()),
                "Total F After": int(g["F Wk3 After Destination Change"].sum()),
                "Min SI After": int(g["Current SI After"].min()),
                "Max SI After": int(g["Current SI After"].max()),
                "Min SS % After": float(g["SS % After"].min()),
                "Max SS % After": float(g["SS % After"].max()),
            }
        )
    return pd.DataFrame(rows)


def autofit(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    autofit(ws)


def write_excel_bytes(detail: pd.DataFrame, summary: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="Optimized Data", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        logic = pd.DataFrame(
            [
                ["Current SI", "= Sum of SI Wk3"],
                ["Current SS%", "= Current SI / Average of SS Wk3"],
                ["F Wk3 After", "Firm cuối cùng của kho sau destination change"],
                ["Net Destination Change", "= F Wk3 After - F Wk3 Original"],
                ["Current SI After", "= Current SI + (F Wk3 After - F Wk3 Original)"],
                ["SS % After", "= Current SI After / Average of SS Wk3"],
                ["Priority mode SI", "Target là cover theo % hướng tới SI = 0"],
                ["Priority mode SS", "Target %SS dùng công thức SI / SS"],
                ["Multi-priority", "Nếu có nhiều kho priority, hệ thống sẽ cấp từng 1 PCS cho nhóm priority trước"],
            ],
            columns=["Field", "Meaning"],
        )
        logic.to_excel(writer, sheet_name="Logic", index=False)

        wb = writer.book
        for name in ["Optimized Data", "Summary", "Logic"]:
            ws = wb[name]
            style_sheet(ws)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header = ws.cell(1, cell.column).value
                    if header in ["Current SS%", "SS % After", "Priority Rule Value"] and isinstance(cell.value, (int, float)) and not math.isinf(cell.value):
                        cell.number_format = "0.0%"

    output.seek(0)
    return output.getvalue()


def write_vendor_template_bytes(uploaded_file) -> Optional[bytes]:
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=INPUT_SHEET_MAIN)
    df.columns = [str(c).strip() for c in df.columns]
    vendor_cols = [c for c in df.columns if "vendor" in str(c).lower()]
    if vendor_cols:
        return None

    template_df = df.copy()
    if VENDOR_FALLBACK_COL not in template_df.columns:
        template_df[VENDOR_FALLBACK_COL] = ""

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        template_df.to_excel(writer, sheet_name=INPUT_SHEET_MAIN, index=False)
        wb = writer.book
        for ws in wb.worksheets:
            style_sheet(ws)

    output.seek(0)
    return output.getvalue()


def run_optimization(input_df: pd.DataFrame, rules: Dict[str, PriorityRule]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    results = []
    for _, group in input_df.groupby("Item", sort=True):
        item_whse = set(group["Whse"].astype(str).tolist())
        item_rules = {wh: rule for wh, rule in rules.items() if wh in item_whse}
        results.append(allocate_item(group.copy(), item_rules))

    detail_full = pd.concat(results, ignore_index=True)
    detail = build_detail_output(detail_full)
    summary = build_summary(detail_full)

    before_all = int(detail_full["F Wk3 Original"].sum())
    after_all = int(detail_full["F Wk3 After Destination Change"].sum())
    if before_all != after_all:
        raise ValueError(f"Sai tổng Firm PO toàn file: before={before_all}, after={after_all}")

    return detail, summary


def init_rule_state(warehouses: List[str]):
    if "rule_state" not in st.session_state:
        st.session_state.rule_state = {}
    current = st.session_state.rule_state
    st.session_state.rule_state = {
        wh: current.get(wh, {"enabled": False, "mode": "SI", "value": 50.0})
        for wh in warehouses
    }


def collect_rules(warehouses: List[str]) -> Dict[str, PriorityRule]:
    rules: Dict[str, PriorityRule] = {}
    for wh in warehouses:
        state = st.session_state.rule_state[wh]
        if state["enabled"]:
            rules[wh] = PriorityRule(
                whse=wh,
                mode=state["mode"],
                value=normalize_pct(state["value"]),
            )
    return rules


def main():
    st.set_page_config(page_title="Destination Change Optimizer", layout="wide")
    st.title("Destination Change Optimizer")
    st.caption("Upload file Excel, chọn priority warehouse, rồi tải file kết quả.")

    with st.expander("Yêu cầu file input", expanded=True):
        st.write(
            "Sheet bắt buộc: `Sheet1` với các cột `Item`, `ProdResourceID`, `Whse`, `F Wk3`, `Sum of SI Wk3`, `Average of SS Wk3`."
        )
        st.write("Công thức SS% đang dùng: `Current SI / Average of SS Wk3`.")

    uploaded_file = st.file_uploader("Upload file Excel", type=["xlsx", "xlsm", "xls"])

    if not uploaded_file:
        st.stop()

    try:
        file_bytes = uploaded_file.getvalue()
        df = load_input(io.BytesIO(file_bytes))
    except Exception as e:
        st.error(f"Không đọc được file input: {e}")
        st.stop()

    st.success(f"Đã đọc file thành công: {uploaded_file.name}")
    st.write(f"Số dòng dữ liệu: {len(df):,}")

    warehouses = sorted(df["Whse"].dropna().astype(str).unique().tolist(), key=lambda x: (len(x), x))
    init_rule_state(warehouses)

    st.subheader("Priority warehouse rules")
    st.write("Bật kho cần ưu tiên, rồi nhập mode và giá trị.")

    cols = st.columns(4)
    cols[0].markdown("**Whse**")
    cols[1].markdown("**Enable**")
    cols[2].markdown("**Mode**")
    cols[3].markdown("**Value**")

    for wh in warehouses:
        row_cols = st.columns(4)
        state = st.session_state.rule_state[wh]
        enabled = row_cols[1].checkbox(
            label=f"enable_{wh}",
            value=state["enabled"],
            key=f"enabled_{wh}",
            label_visibility="collapsed",
        )
        mode = row_cols[2].selectbox(
            label=f"mode_{wh}",
            options=["SI", "SS"],
            index=0 if state["mode"] == "SI" else 1,
            key=f"mode_{wh}",
            label_visibility="collapsed",
        )
        default_value = float(state["value"])
        help_text = "SI: % cover toward SI=0. SS: target SS ratio theo công thức SI/SS."
        value = row_cols[3].number_input(
            label=f"value_{wh}",
            value=default_value,
            step=1.0,
            key=f"value_{wh}",
            label_visibility="collapsed",
            help=help_text,
        )
        row_cols[0].write(wh)
        st.session_state.rule_state[wh] = {
            "enabled": enabled,
            "mode": mode,
            "value": value,
        }

    with st.expander("Rule guide"):
        st.write("- Mode SI: nhập 50 nghĩa là cover 50% khoảng cách về SI = 0.")
        st.write("- Mode SS: nhập 100 nghĩa là target SI = SS, nhập 0 nghĩa là target SI = 0.")
        st.write("- Bạn có thể bật nhiều warehouse priority cùng lúc.")

    preview_cols = [c for c in ["Item", "ProdResourceID", "Whse", "F Wk3", "Current SI", "Average of SS Wk3", "Current SS%", "Vendor"] if c in df.columns]
    st.subheader("Input preview")
    st.dataframe(df[preview_cols].head(50), use_container_width=True)

    if st.button("Run Optimization", type="primary"):
        try:
            rules = collect_rules(warehouses)
            detail, summary = run_optimization(df, rules)
            result_bytes = write_excel_bytes(detail, summary)
            vendor_template_bytes = write_vendor_template_bytes(io.BytesIO(file_bytes))

            st.session_state.result_bytes = result_bytes
            st.session_state.result_name = os.path.splitext(uploaded_file.name)[0] + "_destination_change_optimized.xlsx"
            st.session_state.detail = detail
            st.session_state.summary = summary
            st.session_state.vendor_template_bytes = vendor_template_bytes
            st.session_state.vendor_template_name = os.path.splitext(uploaded_file.name)[0] + "_with_vendor_template.xlsx"
            st.success("Đã xử lý xong.")
        except Exception as e:
            st.error(f"Xử lý thất bại: {e}")

    if "detail" in st.session_state and "summary" in st.session_state:
        st.subheader("Output preview")
        tab1, tab2 = st.tabs(["Optimized Data", "Summary"])
        with tab1:
            st.dataframe(st.session_state.detail, use_container_width=True)
        with tab2:
            st.dataframe(st.session_state.summary, use_container_width=True)

        st.download_button(
            label="Download optimized file",
            data=st.session_state.result_bytes,
            file_name=st.session_state.result_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if st.session_state.vendor_template_bytes is not None:
            st.download_button(
                label="Download vendor template",
                data=st.session_state.vendor_template_bytes,
                file_name=st.session_state.vendor_template_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
